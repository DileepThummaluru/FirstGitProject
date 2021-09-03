import openpyxl


class HomePageData:

    # giving Data manually
    test_HomePageData = [{"firstname": "Dileep", "email": "dil.thu@gmail.com", "password": "dil123", "gender": "Male",
                          "Bday": "05061996"},
                         {"firstname": "Revathi", "email": "revathi.123@gmail.com", "password": "rev123",
                          "gender": "Female", "Bday": "10081996"}]

    # giving data from excel sheet
    @staticmethod
    def get_test_data(test_case1, test_case2):
        dictionary1 = {}
        dictionary2 = {}

        book = openpyxl.load_workbook('C:\\Users\\Dileep Thummaluru\\Documents\\PythonDemo.xlsx')
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case1:
                for j in range(2, sheet.max_column + 1):
                    dictionary1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            elif sheet.cell(row=i, column=1).value == test_case2:
                for j in range(2, sheet.max_column + 1):
                    dictionary2[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dictionary1, dictionary2]
